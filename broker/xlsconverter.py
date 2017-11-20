import errno
import json
import logging
import os
from collections import defaultdict
from itertools import chain

from openpyxl import load_workbook

# these are spreadsheet fields that can be a list
# todo - these should be read out of the json schema at the start
hca_v3_lists = ['seq.lanes']


class SpreadsheetConverter:
    def __init__(self, dry=False, output=None):

        self.logger = logging.getLogger(__name__)
        self.outputDir = output

    # this only works for keys nested to two levels, need something smarter to handle arbitrary
    # depth keys e.g. we support <level1>.<level2> = <value>, where value is either a single value or list
    # this doesn't support <level1>.<level2>.<level3>
    def _key_value_to_nested_object(self, obj, key, value):
        d = value
        if "\"" in str(value) or "||" in str(value) or key in hca_v3_lists:
            # d = map(lambda it: it.strip(' "\''), str(value).split("||"))
            d = str(value).split("||")

        if len(key.split('.')) > 3:
            raise ValueError('We don\'t support keys nested greater than 3 levels, found:' + key)
        for part in reversed(key.split('.')):
            d = {part: d}

        return self._merge_dict(obj, d)

    def _merge_dict(self, dict1, dict2):
        dict3 = defaultdict(list)
        for k, v in chain(dict1.items(), dict2.items()):
            if k in dict3:
                if isinstance(v, dict):
                    dict3[k].update(self._merge_dict(dict3[k], v))
                else:
                    dict3[k].update(v)
            else:
                dict3[k] = v
        return dict3

    # sheets with one or more data rows and properties in row 1
    def _multi_row_to_object_from_sheet(self, type, sheet):
        objs = []
        for row in sheet.iter_rows(row_offset=1, max_row=(sheet.max_row - 1)):
            obj = {}
            has_data = False
            for cell in row:
                if not cell.value and not isinstance(cell.value, (int, float)):
                    continue
                has_data = True
                cell_col = cell.col_idx
                property_value = sheet.cell(row=1, column=cell_col).value

                d = self._key_value_to_nested_object(obj, property_value, cell.value)
                obj.update(d)
            if has_data:
                self.logger.debug(json.dumps(obj))
                objs.append(obj)

        return objs

    # sheets that represent one entity where the properties are in column 0
    def _sheet_to_object(self, type, sheet):
        obj = {}
        for row in sheet.iter_rows():
            if len(row) > 1:
                property_cell = row[0].value
                value_cell = row[1].value
                if value_cell:
                    obj = self._key_value_to_nested_object(obj, property_cell, value_cell)
        self.logger.debug(json.dumps(obj))
        return obj

    def convert(self, spreadsheet_file, output_file=None):
        output = None

        try:
            output = self._process(spreadsheet_file)
        except ValueError as e:
            self.logger.error("Error:" + str(e))
            raise e

        if output_file:
            output.to_file(output_file)

        return output

    def _process(self, path_to_spreadsheet):
        output = SpreadsheetConverterOutput()

        # parse the spreadsheet
        wb = load_workbook(filename=path_to_spreadsheet)
        project_sheet = wb.get_sheet_by_name("project")
        project_pubs_sheet = wb.get_sheet_by_name("project.publications")
        submitter_sheet = wb.get_sheet_by_name("contact.submitter")
        contributor_sheet = wb.get_sheet_by_name("contact.contributors")
        specimen_sheet = wb.get_sheet_by_name("sample.specimen_from_organism")
        specimen_state_sheet = wb.get_sheet_by_name("sample.specimen_from_organism.s")
        donor_sheet = wb.get_sheet_by_name("sample.donor")
        cell_suspension_sheet = wb.get_sheet_by_name("sample.cell_suspension")
        cell_suspension_enrichment_sheet = wb.get_sheet_by_name("sample.cell_suspension.enrichme")
        cell_suspension_well_sheet = wb.get_sheet_by_name("sample.cell_suspension.well")

        organoid_sheet = wb.create_sheet()
        if "sample.organoid" in wb.sheetnames:
            organoid_sheet = wb.get_sheet_by_name("sample.organoid")

        immortalized_CL_sheet = wb.create_sheet()
        if "sample.immortalized_cell_line" in wb.sheetnames:
            immortalized_CL_sheet = wb.get_sheet_by_name("sample.immortalized_cell_line")

        primary_CL_Sheet = wb.create_sheet()
        if "sample.primary_cell_line" in wb.sheetnames:
            primary_CL_Sheet = wb.get_sheet_by_name("sample.primary_cell_line")
        protocol_sheet = wb.get_sheet_by_name("protocols")
        single_cell_sheet = wb.get_sheet_by_name("single_cell")
        single_cell_barcode_sheet = wb.get_sheet_by_name("single_cell.barcode")
        rna_sheet = wb.get_sheet_by_name("rna")
        seq_sheet = wb.get_sheet_by_name("seq")
        seq_barcode_sheet = wb.get_sheet_by_name("seq.barcode")
        files_sheet = wb.get_sheet_by_name("file")

        # convert data in sheets back into dict
        project = self._sheet_to_object("project", project_sheet)
        enrichment = self._multi_row_to_object_from_sheet("enrichment", cell_suspension_enrichment_sheet)
        well = self._multi_row_to_object_from_sheet("well", cell_suspension_well_sheet)
        single_cell = self._multi_row_to_object_from_sheet("single_cell", single_cell_sheet)
        single_cell_barcode = self._multi_row_to_object_from_sheet("barcode", single_cell_barcode_sheet)
        rna = self._multi_row_to_object_from_sheet("rna", rna_sheet)
        seq = self._multi_row_to_object_from_sheet("seq", seq_sheet)
        seq_barcode = self._multi_row_to_object_from_sheet("barcode", seq_barcode_sheet)

        protocols = self._multi_row_to_object_from_sheet("protocol", protocol_sheet)
        donors = self._multi_row_to_object_from_sheet("donor", donor_sheet)
        publications = self._multi_row_to_object_from_sheet("publication", project_pubs_sheet)
        submitters = self._multi_row_to_object_from_sheet("submitter", submitter_sheet)
        contributors = self._multi_row_to_object_from_sheet("contributor", contributor_sheet)
        specimens = self._multi_row_to_object_from_sheet("specimen_from_organism", specimen_sheet)
        specimen_state = self._multi_row_to_object_from_sheet("state_of_specimen", specimen_state_sheet)
        cell_suspension = self._multi_row_to_object_from_sheet("cell_suspension", cell_suspension_sheet)
        organoid = self._multi_row_to_object_from_sheet("organoid", organoid_sheet)
        immortalized_cell_line = self._multi_row_to_object_from_sheet("immortalized_cell_line", immortalized_CL_sheet)
        primary_cell_line = self._multi_row_to_object_from_sheet("primary_cell_line", primary_CL_Sheet)
        files = self._multi_row_to_object_from_sheet("file", files_sheet)

        samples = []
        samples.extend(donors)
        samples.extend(specimens)
        samples.extend(cell_suspension)
        samples.extend(organoid)
        samples.extend(immortalized_cell_line)
        samples.extend(primary_cell_line)

        # post objects to the Ingest API after some basic validation
        if "project_id" not in project:
            raise ValueError('Project must have an id attribute')
        project_id = project["project_id"]

        # embed contact & publication into into project for now
        pubs = []
        for index, publication in enumerate(publications):
            pubs.append(publication)
        project["publications"] = pubs

        subs = []
        for index, submitter in enumerate(submitters):
            subs.append(submitter)
        project["submitters"] = subs

        cont = []
        for index, contributor in enumerate(contributors):
            cont.append(contributor)
        project["contributors"] = cont

        links_list = []

        project["core"] = {"type": "project"}

        output.project = project

        protocol_map = {}
        for index, protocol in enumerate(protocols):
            if "protocol_id" not in protocol:
                raise ValueError('Protocol must have an id attribute')
            protocol_map[protocol["protocol_id"]] = protocol

            output.protocols.append(protocol)

            links_list.append("protocol_" + protocol["protocol_id"] + "-project_" + project_id)

        sample_map = {}
        for index, sample in enumerate(samples):
            if "sample_id" not in sample:
                raise ValueError('Sample must have an id attribute')
            sample_map[sample["sample_id"]] = sample

        # add dependent information to various sample types
        for state in specimen_state:
            if "sample_id" in state:
                sample_map[state["sample_id"]]["specimen_from_organism"]["state_of_specimen"] = state[
                    "state_of_specimen"]

        for e in enrichment:
            if "sample_id" in e:
                if "enrichment" in sample_map[state["sample_id"]]["cell_suspension"]:
                    sample_map[state["sample_id"]]["cell_suspension"]["enrichment"].append(e["enrichment"])
                else:
                    sample_map[state["sample_id"]]["cell_suspension"]["enrichment"] = [e["enrichment"]]

        for w in well:
            if "sample_id" in w:
                sample_map[state["sample_id"]]["cell_suspension"]["well"] = w["well"]

        # create derived_from links between samples
        for index, sample_id in enumerate(sample_map.keys()):
            sample = sample_map[sample_id]
            if "derived_from" in sample:
                if sample["derived_from"] not in sample_map.keys():
                    raise ValueError('Sample ' + sample_id + ' references another sample ' + sample[
                        "derived_from"] + ' that isn\'t in the donor worksheet')
            sample_protocols = []
            if "protocol_ids" in sample:
                for sampleProtocolId in sample["protocol_ids"]:
                    if sampleProtocolId not in protocol_map:
                        raise ValueError('Sample ' + sample[
                            "sample_id"] + ' references a protocol ' + sampleProtocolId + ' that isn\'t in the protocol worksheet')
                sample_protocols = sample["protocol_ids"]
                sample["protocol_ids"]

            sample["core"] = {"type": "sample"}

            output.samples.append(sample)

            if "derived_from" in sample_map[sample_id]:
                links_list.append(
                    "sample_" + sample_id + "-derivedFromSamples_" + sample_map[sample_id]["derived_from"])

            if sample_protocols:
                for sampleProtocolId in sample_protocols:
                    links_list.append("sample_" + sample_id + "-protocol_" + sampleProtocolId)

        # build the assay map from the different types of assay infromation
        assay_map = {}

        for index, s in enumerate(seq):
            if "assay_id" in s:
                id = s["assay_id"]
                del s["assay_id"]
                assay_map[id]["seq"] = s["seq"]
            else:
                seq_obj = s["seq"]

        for sb in seq_barcode:
            if "assay_id" in sb:
                id = sb["assay_id"]
                del sb["assay_id"]
                assay_map[id]["seq"]["umi_barcode"] = sb["umi_barcode"]
            else:
                seq_obj["umi_barcode"] = sb["umi_barcode"]

        for index, sc in enumerate(single_cell):
            if "assay_id" in sc:
                id = sc["assay_id"]
                del sc["assay_id"]
                assay_map[id]["single_cell"] = sc["single_cell"]
            else:
                sc_obj = sc["single_cell"]

        for scb in single_cell_barcode:
            if "assay_id" in scb:
                id = scb["assay_id"]
                del scb["assay_id"]
                assay_map[id]["single_cell"]["barcode"] = scb["barcode"]
            else:
                sc_obj["barcode"] = scb["barcode"]

        for index, r in enumerate(rna):
            if "assay_id" in r:
                id = r["assay_id"]
                del r["assay_id"]
                assay_map[id]["rna"] = r["rna"]
            else:
                rnaObj = r["rna"]

        files_map = {}
        for index, file in enumerate(files):
            if "filename" not in file:
                raise ValueError('Files must have a name')
            if "assay_id" not in file:
                raise ValueError('Files must be linked to an assay')
            assay = file["assay_id"]
            seq_file = file["seq"]
            sample = file["sample_id"]

            files_map[file["filename"]] = file

            if assay not in assay_map:
                assay_map[assay] = {}
                assay_map[assay]["files"] = []
                assay_map[assay]["rna"] = rnaObj.copy()
                assay_map[assay]["single_cell"] = sc_obj.copy()
                assay_map[assay]["seq"] = seq_obj.copy()
                assay_map[assay]["seq"]["lanes"] = []
                assay_map[assay]["sample_id"] = sample
                assay_map[assay]["assay_id"] = assay
            elif "rna" not in assay_map[assay]:
                assay_map[assay]["rna"] = rnaObj.copy()
            elif "single_cell" not in assay_map[assay].keys():
                assay_map[assay]["single_cell"] = sc_obj.copy()
            elif "seq" not in assay_map[assay]:
                assay_map[assay]["seq"] = seq_obj.copy()
                assay_map[assay]["seq"]["lanes"] = []
            elif "sample_id" not in assay_map[assay]:
                assay_map[assay]["sample_id"] = sample
            elif "files" not in assay_map[assay]:
                assay_map[assay]["files"] = []
            elif "assay_id" not in assay_map[assay]:
                assay_map[assay]["assay_id"] = assay

            assay_map[assay]["files"].append(file["filename"])

            if "lanes" in seq_file:
                if "number" in seq_file["lanes"]:
                    added = False
                    if len(assay_map[assay]["seq"]["lanes"]) > 0:
                        for lane in assay_map[assay]["seq"]["lanes"]:
                            if lane["number"] == seq_file["lanes"]["number"]:
                                if "run" in seq_file["lanes"]:
                                    lane[seq_file["lanes"]["run"]] = file["filename"]
                                    added = True
                    if not added:
                        if "run" in seq_file["lanes"]:
                            assay_map[assay]["seq"]["lanes"].append({"number": seq_file["lanes"]["number"],
                                                                     seq_file["lanes"]["run"]: file["filename"]})
                else:
                    if "run" in seq_file["lanes"]:
                        assay_map[assay]["seq"]["lanes"].append({seq_file["lanes"]["run"]: file["filename"]})

            if "insdc_experiment" in seq_file:
                assay_map[assay]["seq"]["insdc_experiment"] = seq_file["insdc_experiment"]

            if "insdc_run" in seq_file:
                assay_map[assay]["seq"]["insdc_run"] = seq_file["insdc_run"]

            file["core"] = {"type": "file"}

            output.files.append(file)

            if sample in sample_map:
                links_list.append("file_" + file["filename"] + "-sample_" + sample)

        for index, assay in enumerate(assay_map.values()):
            if "assay_id" not in assay:
                raise ValueError('Each assay must have an id attribute' + str(assay))
            if "files" not in assay:
                raise ValueError('Each assay must list associated files using the files attribute')
            else:
                for file in assay["files"]:
                    if file not in files_map:
                        raise ValueError('Assay references file ' + file + ' that isn\'t defined in the files sheet')
            files = assay["files"]
            # del assay["files"]

            if "sample_id" not in assay:
                raise ValueError("Every assay must reference a sample using the sample_id attribute")
            elif assay["sample_id"] not in sample_map:
                raise ValueError(
                    'An assay references a sample ' + assay["sample_id"] + ' that isn\'t in the samples worksheet')
            samples = assay["sample_id"]

            assay["core"] = {"type": "assay"}

            output.assays.append(assay)

            if samples in sample_map:
                links_list.append("assay_" + assay["assay_id"] + "-sample_" + samples)

            for file in files:
                links_list.append("assay_" + assay["assay_id"] + "-file_" + file)

        output.links = links_list
        self.logger.info("All done!")
        wb.close()
        return output


class SpreadsheetConverterOutput:
    def __init__(self):
        self.project = None
        self.protocols = []
        self.donors = []
        self.samples = []
        self.files = []
        self.assays = []
        self.links = []

    def to_file(self, filename):
        if filename:
            if not os.path.exists(os.path.dirname(filename)):
                try:
                    os.makedirs(os.path.dirname(filename))
                except OSError as exc:
                    if exc.errno != errno.EEXIST:
                        raise

            with open(filename, "w") as f:
                f.write(json.dumps(vars(self), indent=4))
